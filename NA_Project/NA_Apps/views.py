# NA_Apps/views.py
from __future__ import annotations

from django.shortcuts import render, redirect, get_object_or_404
from django.utils     import timezone
from django.contrib   import messages
from django.db        import transaction
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
import xlsxwriter
from io import BytesIO
import json
from django.views.decorators.http import require_http_methods, require_POST, require_GET
from django.template.loader import render_to_string
from django.views.decorators.csrf import ensure_csrf_cookie
from datetime import datetime

from .models          import Device, Log
from .forms           import UploadFileForm
from .mikrotik_api_bulk import open_api, exec_cli
from .filters         import SegmentationFilter
from segmentation.models import Segmentation

import ipaddress, pandas as pd
import socket
import paramiko
import time


def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
        
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome back, {username}!')
            return redirect('home')
        else:
            messages.error(request, 'Invalid username or password.')
            
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    messages.info(request, 'You have been logged out successfully.')
    return redirect('login')

# Protect all views with login_required
@login_required(login_url='login')
def home(request):
    return render(request, "home.html", {
        "all_device": Device.objects.count(),
        "mikrotik_device": Device.objects.filter(vendor__iexact="mikrotik").count(),
        "logs": Log.objects.all()[:10],
    })


# ╭────────────────────── LOG VIEW & MASS DELETE ───────────────────╮
@login_required(login_url='login')
def mass_delete_log(request):
    log_type = request.GET.get('type', 'config')  # Default to config logs
    
    if request.method == "POST":
        if log_type == 'config':
            # Delete configuration related logs
            Log.objects.filter(action__in=["Configure", "Connect API"]).delete()
            messages.success(request, "Configuration logs deleted successfully.")
            return redirect("log")
        elif log_type == 'verify':
            # Delete verify logs
            Log.objects.filter(action="Verify").delete()
            messages.success(request, "Verify logs deleted successfully.")
            return redirect("log")
        else:
            # Delete device management logs
            Log.objects.filter(action="Add Device").delete()
            messages.success(request, "Device logs deleted successfully.")
            return redirect("device_logs")

    # Filter logs based on type
    if log_type == 'config':
        logs = Log.objects.filter(action__in=["Configure", "Connect API"])
        title = "Clear Configuration Logs"
        return_url = 'log'
    elif log_type == 'verify':
        logs = Log.objects.filter(action="Verify")
        title = "Clear Verify Logs"
        return_url = 'log'
    else:
        logs = Log.objects.filter(action="Add Device")
        title = "Clear Device Logs"
        return_url = 'device_logs'

    # Get all devices for hostname lookup
    devices = Device.objects.all()

    return render(request, "mass_delete_log.html", {
        "logs": logs,
        "devices": devices,
        "title": title,
        "return_url": return_url,
        "log_type": log_type
    })


@login_required(login_url='login')
def log(request):
    # Base query - include configuration and verify logs
    logs = Log.objects.all()
    
    # Get filter parameters
    hostname = request.GET.get('hostname', '')
    status = request.GET.get('status', '')
    
    # Apply filters if provided
    if hostname:
        logs = logs.filter(target__icontains=hostname)
    
    # Create a copy of the queryset for counting before status filtering
    filtered_logs = logs
    
    # Apply status filter if provided
    if status:
        logs = logs.filter(status=status)
    
    # Get unique hostnames and statuses for filter dropdowns
    all_hostnames = Log.objects.values_list('target', flat=True).distinct()
    all_statuses = [choice[0] for choice in Log.STATUS_CHOICES]
    
    # Count success and error logs based on the filtered queryset
    if status:
        # If status filter is applied, show filtered counts
        success_count = logs.filter(status='Success').count()
        error_count = logs.filter(status='Error').count()
    else:
        # If no status filter, show total counts
        success_count = filtered_logs.filter(status='Success').count()
        error_count = filtered_logs.filter(status='Error').count()
    
    context = {
        'logs': logs,
        'all_hostnames': all_hostnames,
        'all_statuses': all_statuses,
        'selected_hostname': hostname,
        'selected_status': status,
        'success_count': success_count,
        'error_count': error_count,
    }
    
    # If it's an HTMX request, return only the logs partial
    if request.headers.get('HX-Request'):
        return render(request, 'partials/log_list.html', context)
    
    return render(request, 'log.html', context)

@login_required(login_url='login')
@require_POST
@ensure_csrf_cookie
def clear_logs(request):
    """Clear all logs"""
    try:
        # Delete all logs
        count = Log.objects.all().count()
        Log.objects.all().delete()
        
        # Return success response
        return JsonResponse({
            'success': True,
            'message': f'Successfully cleared {count} logs'
        })
    except Exception as e:
        # Return error response with detailed message
        return JsonResponse({
            'success': False,
            'error': f'Failed to clear logs: {str(e)}'
        }, status=500)


# ╭────────────────── MASS-CONFIG (WRITE) VIA API ───────────────────╮
@login_required(login_url='login')
def configure(request):
    if request.method != "POST":
        devices = Device.objects.all()
        
        # Apply device category filter
        device_category = request.GET.get('device_category', '')
        if device_category:
            devices = devices.filter(device_category=device_category)
        
        # Get device categories for filter dropdown
        device_categories = Device.DEVICE_CATEGORY_CHOICES
        
        return render(request, "config.html", {
            "devices": devices,
            "device_category": device_category,
            "device_categories": device_categories,
            "mode": "Configure",
        })

    cmds = [ln.strip() for ln in
            request.POST.get("mikrotik_command", "").splitlines()
            if ln.strip() and not ln.lstrip().startswith("#")]

    for dev_id in request.POST.getlist("device"):
        dev = get_object_or_404(Device, pk=dev_id)

        if dev.vendor.lower() != "mikrotik":
            Log.objects.create(
                target=dev.ip_address,
                action="Configure",
                status="Error",
                time=timezone.now(),
                messages="Unsupported vendor for API",
                command="\n".join(cmds)
            )
            continue

        api, pool = open_api(dev.ip_address,
                            dev.api_port or 8728,
                            dev.username,
                            dev.password)
        if api is None:
            continue

        try:
            for raw in cmds:
                exec_cli(api, raw)

            Log.objects.create(
                target=dev.ip_address,
                action="Configure",
                status="Success",
                time=timezone.now(),
                messages=f"Configuration applied successfully via port {dev.api_port or 8728}",
                command="\n".join(cmds)
            )
        except Exception as e:
            Log.objects.create(
                target=dev.ip_address,
                action="Configure",
                status="Error",
                time=timezone.now(),
                messages=str(e),
                command="\n".join(cmds)
            )
        finally:
            try:
                pool.disconnect()
            except:
                pass

    return redirect('log')
# ╰──────────────────────────────────────────────────────────────────╯


# ╭────────────────── VERIFY (READ-ONLY) VIA API ────────────────────╮
@login_required(login_url='login')
def verify_config(request):
    if request.method != "POST":
        devices = Device.objects.all()
        
        # Apply device category filter
        device_category = request.GET.get('device_category', '')
        if device_category:
            devices = devices.filter(device_category=device_category)
        
        # Get device categories for filter dropdown
        device_categories = Device.DEVICE_CATEGORY_CHOICES
        
        return render(request, "config.html", {
            "devices": devices,
            "device_category": device_category,
            "device_categories": device_categories,
            "mode": "Verify",
        })

    cmds = [ln.strip() for ln in
            request.POST.get("mikrotik_command", "").splitlines()
            if ln.strip()]

    results = []

    for dev_id in request.POST.getlist("device"):
        dev = get_object_or_404(Device, pk=dev_id)

        if dev.vendor.lower() != "mikrotik":
            Log.objects.create(
                target=dev.ip_address,
                action="Verify",
                status="Error",
                time=timezone.now(),
                messages="Unsupported vendor (only MikroTik)",
                command="\n".join(cmds)
            )
            continue

        api, pool = open_api(dev.ip_address,
                            dev.api_port or 8728,
                            dev.username,
                            dev.password)
        if api is None:
            Log.objects.create(
                target=dev.ip_address,
                action="Verify",
                status="Error",
                time=timezone.now(),
                messages=f"Failed to connect to device via API (Port: {dev.api_port or 8728})",
                command="\n".join(cmds)
            )
            continue

        try:
            device_output = []
            for raw in cmds:
                try:
                    # Execute command and get output
                    output = exec_cli(api, raw)
                    
                    # Format the output for display
                    formatted_output = f"\n[{dev.ip_address}:{dev.api_port or 8728}] Command: {raw}\n"
                    formatted_output += "-" * 60 + "\n"
                    
                    if isinstance(output, list):
                        formatted_output += "\n".join(str(item) for item in output)
                    else:
                        formatted_output += str(output)
                    
                    formatted_output += "\n" + "-" * 60 + "\n"
                    device_output.append(formatted_output)

                    # Create success log with command output
                    Log.objects.create(
                        target=dev.ip_address,
                        action="Verify",
                        status="Success",
                        time=timezone.now(),
                        messages=f"Command executed successfully",
                        command=raw,
                        output=str(output)
                    )
                except Exception as cmd_error:
                    error_msg = f"Error executing '{raw}': {str(cmd_error)}"
                    device_output.append(f"\n[{dev.ip_address}:{dev.api_port or 8728}] {error_msg}\n")
                    
                    # Create error log for failed command
                    Log.objects.create(
                        target=dev.ip_address,
                        action="Verify",
                        status="Error",
                        time=timezone.now(),
                        messages=error_msg,
                        command=raw
                    )

            # Join all outputs for this device
            results.extend(device_output)

        except Exception as e:
            error_msg = f"Connection Error: {str(e)}"
            results.append(f"\n[{dev.ip_address}:{dev.api_port or 8728}] {error_msg}\n")
            
            Log.objects.create(
                target=dev.ip_address,
                action="Verify",
                status="Error",
                time=timezone.now(),
                messages=error_msg,
                command="\n".join(cmds)
            )
        finally:
            try:
                pool.disconnect()
            except:
                pass

    return redirect('log')

@login_required(login_url='login')
def verify_config_ssh(request):
    if request.method != "POST":
        return render(request, "config.html", {
            "devices": Device.objects.all(),
            "mode": "Verify",
        })

    cmds = request.POST['mikrotik_command'].splitlines()

    for dev_id in request.POST.getlist('device'):
        dev = get_object_or_404(Device, pk=dev_id)

        sock = _open_ssh_socket(dev.ip_address, dev.ssh_port or 22,
                              timeout=5, retries=2, delay=2)
        if not sock:
            Log.objects.create(
                target=dev.ip_address,
                action="Verify",
                status="Error",
                time=timezone.now(),
                messages=f"Connection failed: {dev.ip_address}:{dev.ssh_port or 22}",
                command="\n".join(cmds)
            )
            continue

        try:
            transport = paramiko.Transport(sock)
            sec = transport.get_security_options()
            sec.digests = ['hmac-sha1', 'hmac-md5']

            transport.start_client()
            transport.auth_password(dev.username, dev.password)

            ssh = paramiko.SSHClient()
            ssh._transport = transport

            output = []
            for cmd in cmds:
                stdin, stdout, stderr = ssh.exec_command(cmd)
                output.extend([
                    f"\nCommand: {cmd}",
                    stdout.read().decode(),
                    stderr.read().decode()
                ])

            output_text = "\n".join(output)
            Log.objects.create(
                target=dev.ip_address,
                action="Verify",
                status="Success",
                time=timezone.now(),
                messages="Commands executed successfully",
                command="\n".join(cmds),
                output=output_text
            )
        except Exception as e:
            Log.objects.create(
                target=dev.ip_address,
                action="Verify",
                status="Error",
                time=timezone.now(),
                messages=str(e),
                command="\n".join(cmds)
            )
        finally:
            try: transport.close()
            except: pass
            try: sock.close()
            except: pass

    return redirect('log')

@login_required(login_url='login')
def configure_ssh(request):
    if request.method != "POST":
        return render(request, "config.html", {
            "devices": Device.objects.all(),
            "mode": "Configure",
        })

    cmds = request.POST['mikrotik_command'].splitlines()

    for dev_id in request.POST.getlist('device'):
        dev = get_object_or_404(Device, pk=dev_id)

        sock = _open_ssh_socket(dev.ip_address, dev.ssh_port or 22,
                              timeout=5, retries=2, delay=2)
        if not sock:
            Log.objects.create(
                target=dev.ip_address,
                action="Configure",
                status="Error",
                time=timezone.now(),
                messages=f"Connection failed: {dev.ip_address}:{dev.ssh_port or 22}",
                command="\n".join(cmds)
            )
            continue

        try:
            transport = paramiko.Transport(sock)
            sec = transport.get_security_options()
            sec.digests = ['hmac-sha1', 'hmac-md5']

            transport.start_client()
            transport.auth_password(dev.username, dev.password)

            ssh = paramiko.SSHClient()
            ssh._transport = transport

            output = []
            for cmd in cmds:
                stdin, stdout, stderr = ssh.exec_command(cmd)
                output.extend([stdout.read().decode(), stderr.read().decode()])

            Log.objects.create(
                target=dev.ip_address,
                action="Configure",
                status="Success",
                time=timezone.now(),
                messages="Configuration applied successfully",
                command="\n".join(cmds)
            )
        except Exception as e:
            Log.objects.create(
                target=dev.ip_address,
                action="Configure",
                status="Error",
                time=timezone.now(),
                messages=str(e),
                command="\n".join(cmds)
            )
        finally:
            try: transport.close()
            except: pass
            try: sock.close()
            except: pass

    return redirect('log')

# ╭──────────────────────── HOME & DEVICE CRUD ──────────────────────╮
@login_required(login_url='login')
def devices(request):
    # Get all devices
    devices = Device.objects.all()
    
    # Ensure both Distribution and Customer segments exist
    distribution = Segmentation.objects.filter(segmentation_type='distribution').first()
    customer = Segmentation.objects.filter(segmentation_type='customer').first()
    
    # Create segments if they don't exist
    if not distribution:
        distribution = Segmentation.objects.create(
            name='Distribution',
            segmentation_type='distribution'
        )
    
    if not customer:
        customer = Segmentation.objects.create(
            name='Customer',
            segmentation_type='customer'
        )
    
    # Remove any duplicate segments but keep the ones being used by devices
    used_distribution_ids = Device.objects.filter(segmentation__segmentation_type='distribution').values_list('segmentation_id', flat=True)
    used_customer_ids = Device.objects.filter(segmentation__segmentation_type='customer').values_list('segmentation_id', flat=True)
    
    # Keep the first distribution segment and any that are being used
    Segmentation.objects.filter(
        segmentation_type='distribution'
    ).exclude(
        id__in=list(used_distribution_ids) + [distribution.id]
    ).delete()
    
    # Keep the first customer segment and any that are being used
    Segmentation.objects.filter(
        segmentation_type='customer'
    ).exclude(
        id__in=list(used_customer_ids) + [customer.id]
    ).delete()
    
    # Get device categories for the dropdown
    device_categories = Device.DEVICE_CATEGORY_CHOICES
    
    # Get final list of segmentations, ordered by type to ensure Distribution appears first
    segmentations = Segmentation.objects.order_by('segmentation_type')
    
    return render(request, "devices.html", {
        "all_device": devices,
        "device_categories": device_categories,
        "segmentations": segmentations,
        "total_devices": devices.count(),
    })

@login_required(login_url='login')
@require_GET
def device_json(request, pk):
    """Get device data in JSON format for modal"""
    try:
        device = get_object_or_404(Device, pk=pk)
        data = {
            'id': device.id,
            'ip_address': device.ip_address,
            'hostname': device.hostname,
            'username': device.username,
            'password': device.password,
            'api_port': device.api_port,
            'ssh_port': device.ssh_port,
            'device_category': device.device_category,
            'segmentation': {
                'id': device.segmentation.id,
                'name': device.segmentation.name,
                'segmentation_type': device.segmentation.segmentation_type
            } if device.segmentation else None
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required(login_url='login')
@require_POST
def device_save(request, pk):
    try:
        data = json.loads(request.body)
        device = get_object_or_404(Device, pk=pk)
        
        # Update basic fields
        device.ip_address = data['ip_address']
        device.hostname = data['hostname']
        device.username = data['username']
        if data.get('password'):  # Only update password if provided
            device.password = data['password']
        device.vendor = data['vendor']
        device.api_port = data['api_port']
        device.ssh_port = data['ssh_port']
        device.device_category = data['device_category']
        
        # Update segmentation
        if 'segmentation_id' in data:
            try:
                # First try to get existing segmentation
                segmentation = Segmentation.objects.get(id=data['segmentation_id'])
            except Segmentation.DoesNotExist:
                # If not found, create new one based on type
                segmentation_type = data.get('segmentation_type', '').lower()
                if segmentation_type in ['distribution', 'customer']:
                    segmentation = Segmentation.objects.create(
                        name=segmentation_type.capitalize(),
                        segmentation_type=segmentation_type
                    )
                else:
                    raise ValueError(f"Invalid segmentation type: {segmentation_type}")
            
            device.segmentation = segmentation
        
        device.save()
        
        # Create log entry
        Log.objects.create(
            target=device.ip_address,
            action="Edit Device",
            status="Success",
            time=timezone.now(),
            messages=f"Device {device.hostname} updated successfully"
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Device updated successfully',
            'device': {
                'id': device.id,
                'segmentation': {
                    'id': device.segmentation.id,
                    'type': device.segmentation.segmentation_type,
                    'display': device.segmentation.get_segmentation_type_display()
                } if device.segmentation else None
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required(login_url='login')
@require_http_methods(["DELETE"])
def device_delete(request, pk):
    """Delete a device"""
    try:
        device = get_object_or_404(Device, pk=pk)
        ip_address = device.ip_address
        hostname = device.hostname
        
        device.delete()
        
        # Log the action
        Log.objects.create(
            target=ip_address,
            action="Configure",
            status="Success",
            messages=f"Device deleted: {hostname or ip_address}",
            time=timezone.now()
        )
        
        return JsonResponse({'success': True, 'message': 'Device deleted successfully'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required(login_url='login')
def bulk_delete_devices(request):
    """Delete multiple devices at once"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    try:
        data = json.loads(request.body)
        device_ids = data.get('device_ids', [])
        
        if not device_ids:
            return JsonResponse({'error': 'No devices selected'}, status=400)
            
        devices = Device.objects.filter(id__in=device_ids)
        count = devices.count()
        
        # Get device info for logging before deletion
        device_info = [f"{device.hostname or device.ip_address}" for device in devices]
        
        # Delete the devices
        devices.delete()
        
        # Log the action
        Log.objects.create(
            target="Bulk Delete",
            action="Configure",
            status="Success",
            messages=f"Deleted {count} devices: {', '.join(device_info)}",
            time=timezone.now()
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully deleted {count} devices'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required(login_url='login')
@require_POST
def device_add(request):
    """Add a new device"""
    try:
        data = json.loads(request.body)
        
        # Validate required fields
        required_fields = ['ip_address', 'username', 'password']
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return JsonResponse({
                'error': f'Missing required fields: {", ".join(missing_fields)}'
            }, status=400)
        
        # Create device data dictionary
        device_data = {
            'ip_address': data.get('ip_address'),
            'hostname': data.get('hostname', ''),
            'username': data.get('username'),
            'password': data.get('password'),
            'api_port': data.get('api_port', 8728),
            'ssh_port': data.get('ssh_port', 22),
            'vendor': data.get('vendor', 'mikrotik'),
            'device_category': data.get('device_category', 'router_end_point')
        }
        
        # Handle segmentation
        segmentation_type = data.get('segmentation_type')
        if segmentation_type:
            try:
                # Convert to lowercase for case-insensitive comparison
                segmentation_type = segmentation_type.lower()
                segmentation = Segmentation.objects.filter(segmentation_type=segmentation_type).first()
                if not segmentation:
                    # Create new segmentation if it doesn't exist
                    segmentation = Segmentation.objects.create(
                        name=segmentation_type.capitalize(),
                        segmentation_type=segmentation_type
                    )
                device_data['segmentation'] = segmentation
            except Exception as e:
                return JsonResponse({
                    'error': f'Error setting segmentation: {str(e)}'
                }, status=400)
        
        try:
            device = Device.objects.create(**device_data)
        except IntegrityError:
            return JsonResponse({
                'error': f'A device with IP address {data.get("ip_address")} already exists'
            }, status=400)
        
        # Log the action
        Log.objects.create(
            target=device.ip_address,
            action="Add Device",
            status="Success",
            messages=f"Device added: {device.hostname or device.ip_address}",
            time=timezone.now()
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Device added successfully',
            'id': device.id
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=400)


def mass_add_device(request):
    # Import required modules
    import os
    import tempfile
    import pandas as pd
    from django.db import transaction
    from django.utils import timezone
    from segmentation.models import Segmentation
    
    # Initialize context
    context = {
        'title': 'Mass Add Devices',
        'mode': 'upload'
    }
    
    # Handle file upload
    if request.method == "POST" and "file" in request.FILES:
        uploaded_file = request.FILES['file']
        
        # Save file temporarily
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, uploaded_file.name)
        
        with open(temp_path, 'wb+') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)
        
        # Store file path in session
        request.session['uploaded_file_path'] = temp_path
        request.session['uploaded_file_name'] = uploaded_file.name
        
        context.update({
            'mode': 'process',
            'uploaded_file': uploaded_file.name
        })
    
    # Process uploaded file
    elif request.method == "POST" and "process" in request.POST:
        file_path = request.session.get('uploaded_file_path')
        if file_path:
            try:
                df = pd.read_excel(file_path)
                success_count = 0
                error_count = 0
                
                # Expected column names from template
                expected_columns = ['ip', 'hostname', 'username', 'password', 'vendor', 'ssh port', 'api port', 'segment', 'device category']
                
                # Convert column names to lowercase for case-insensitive matching
                df.columns = [col.lower().strip() for col in df.columns]
                
                # Validate columns
                missing_columns = [col for col in expected_columns if col not in df.columns]
                if missing_columns:
                    raise ValueError(f"Missing required columns: {', '.join(missing_columns)}")
                
                with transaction.atomic():
                    for _, row in df.iterrows():
                        try:
                            # Get required fields
                            ip_address = str(row['ip']).strip() if not pd.isna(row['ip']) else None
                            hostname = str(row['hostname']).strip() if not pd.isna(row['hostname']) else None
                            username = str(row['username']).strip() if not pd.isna(row['username']) else None
                            password = str(row['password']).strip() if not pd.isna(row['password']) else None
                            vendor = str(row['vendor']).strip() if not pd.isna(row['vendor']) else 'Mikrotik'
                            
                            # Get port values with defaults
                            try:
                                ssh_port = int(row['ssh port']) if not pd.isna(row['ssh port']) else 22
                            except:
                                ssh_port = 22
                                
                            try:
                                api_port = int(row['api port']) if not pd.isna(row['api port']) else 8728
                            except:
                                api_port = 8728
                            
                            # Get segment and device category
                            segment = str(row['segment']).strip().lower() if not pd.isna(row['segment']) else None
                            device_category = str(row['device category']).strip() if not pd.isna(row['device category']) else None
                            
                            # Map device category to choices
                            category_mapping = {
                                'router end point': 'router_end_point',
                                'access switch': 'access_switch',
                                'router failover': 'router_failover',
                                'radio bts': 'radio_bts',
                                'radio station': 'radio_station',
                                'router bridging': 'router_bridging'
                            }
                            
                            # Convert device category to system format
                            if device_category:
                                device_category = device_category.lower()
                                device_category = category_mapping.get(device_category, device_category)
                            
                            # Validate required fields
                            if not all([ip_address, username, password]):
                                raise ValueError("Missing required fields (IP, username, or password)")
                            
                            # Create device instance
                            device_data = {
                                'hostname': hostname if hostname else f"Device-{ip_address}",
                                'username': username,
                                'password': password,
                                'vendor': vendor,
                                'api_port': api_port,
                                'ssh_port': ssh_port,
                                'device_category': device_category,
                            }
                            
                            # Handle segmentation
                            if segment:
                                try:
                                    segmentation = Segmentation.objects.get(segmentation_type=segment)
                                    device_data['segmentation'] = segmentation
                                except Segmentation.DoesNotExist:
                                    # Create new segmentation if it doesn't exist
                                    segmentation = Segmentation.objects.create(
                                        name=segment.capitalize(),
                                        segmentation_type=segment
                                    )
                                    device_data['segmentation'] = segmentation
                                except Segmentation.MultipleObjectsReturned:
                                    # If multiple segments exist, use the first one
                                    segmentation = Segmentation.objects.filter(segmentation_type=segment).first()
                                    device_data['segmentation'] = segmentation
                            
                            # Create or update device
                            device, created = Device.objects.update_or_create(
                                ip_address=ip_address,
                                defaults=device_data
                            )
                            
                            # Log success
                            Log.objects.create(
                                target=ip_address,
                                action="Add Device",
                                status="Success",
                                time=timezone.now(),
                                messages=f"Device {'added' if created else 'updated'} successfully"
                            )
                            success_count += 1
                            
                        except Exception as e:
                            error_msg = str(e)
                            # Get IP for logging, fallback to "unknown" if not available
                            ip = ip_address if 'ip_address' in locals() else row.get('ip', 'unknown')
                            
                            Log.objects.create(
                                target=ip,
                                action="Add Device",
                                status="Failed",
                                time=timezone.now(),
                                messages=error_msg
                            )
                            error_count += 1
                
                # Clean up the temporary file
                if os.path.exists(file_path):
                    os.remove(file_path)
                    
                # Clear session data
                if 'uploaded_file_path' in request.session:
                    del request.session['uploaded_file_path']
                if 'uploaded_file_name' in request.session:
                    del request.session['uploaded_file_name']
                
                from django.contrib import messages
                messages.success(request, f"Processing complete! {success_count} devices added successfully, {error_count} failed.")
                return redirect("log")
                
            except Exception as e:
                context.update({
                    "error": f"Error processing file: {str(e)}"
                })
        else:
            context.update({
                "error": "No file was uploaded. Please upload a file first."
            })
    
    # Pass the uploaded filename to the template if it exists in session
    if 'uploaded_file_name' in request.session:
        context['uploaded_file'] = request.session['uploaded_file_name']
    
    return render(request, "upload.html", context)


def handle_uploaded_file(f):
    """Save uploaded file to a temporary location and return the path"""
    import os
    import tempfile
    from django.conf import settings
    
    # Create temp directory if it doesn't exist
    temp_dir = os.path.join(settings.BASE_DIR, 'temp_uploads')
    os.makedirs(temp_dir, exist_ok=True)
    
    # Create a unique filename
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', dir=temp_dir)
    
    # Save uploaded file to temp location
    with open(temp_file.name, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    
    return temp_file.name


def hapus_perangkat(request):
    if request.method == "POST":
        Device.objects.filter(id__in=request.POST.getlist("devices")).delete()
        return redirect("home")

    return render(request, "hapus_perangkat.html",
                  {"devices": Device.objects.all()})

@login_required(login_url='login')
def edit_devices(request):
    if request.method == "POST":
        device_ids = request.POST.getlist('devices')
        hostname = request.POST.get('hostname')
        username = request.POST.get('username')
        password = request.POST.get('password')
        ssh_port = request.POST.get('ssh_port')
        api_port = request.POST.get('api_port')
        
        # Update only the fields that were filled
        update_fields = {}
        if hostname: update_fields['hostname'] = hostname
        if username: update_fields['username'] = username
        if password: update_fields['password'] = password
        if ssh_port: 
            update_fields['ssh_port'] = int(ssh_port)  # Convert to integer
        if api_port: 
            update_fields['api_port'] = int(api_port)  # Convert to integer
        
        if update_fields:
            updated = Device.objects.filter(id__in=device_ids).update(**update_fields)
            messages.success(request, f"Successfully updated {len(device_ids)} device(s)")
        else:
            messages.warning(request, "No fields were selected for update")
        
        return redirect('devices')
        
    return redirect('devices')

def _open_ssh_socket(host, port, timeout=5, retries=2, delay=2):
    """
    Buka koneksi TCP ke host:port dengan retry dan timeout.
    Kembalikan socket jika berhasil, atau None jika gagal.
    """
    for attempt in range(1, retries + 1):
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(timeout)
            sock.connect((host, port))
            return sock
        except socket.timeout:
            print(f"[Attempt {attempt}] Timeout {timeout}s ke {host}:{port}")
        except socket.error as e:
            if getattr(e, 'errno', None) == 10060:
                print(f"[Attempt {attempt}] WinError10060: {host}:{port} tidak terjawab")
            else:
                print(f"[Attempt {attempt}] Socket error: {e}")
        time.sleep(delay)
    return None


def download_device_template(request):
    """Generate a template Excel file with updated columns including Segment."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Device Template'

    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    centered = Alignment(horizontal='center')

    # Define headers - updated to include Segment column
    headers = ['IP', 'Hostname', 'Username', 'Password', 'Vendor', 'SSH Port', 'API Port', 'Segment', 'Device Category']

    # Apply headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = centered

    # Add sample data (one row)
    sample_data = [
        '192.168.1.1', 'RouterName', 'admin', 'password123', 'mikrotik', '22', '8728', 'Distribution', 'router_end_point'
    ]

    for col_num, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.border = thin_border
        cell.alignment = centered

    # Auto-size columns
    for col_num, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15

    # Add data validation for device category
    from openpyxl.worksheet.datavalidation import DataValidation

    # Get choices from model
    from NA_Apps.models import Device
    choices = [choice[0] for choice in Device.DEVICE_CATEGORY_CHOICES]

    # Device category validation
    dv_category = DataValidation(
        type='list',
        formula1=f'"{",".join(choices)}"',
        allow_blank=True
    )
    dv_category.error = 'Please select a valid device category'
    dv_category.errorTitle = 'Invalid Category'

    ws.add_data_validation(dv_category)
    dv_category.add('I2:I1000')  # Column I for Device Category

    # Segment validation
    dv_segment = DataValidation(
        type='list',
        formula1='"Distribution,Customer"',
        allow_blank=True
    )
    dv_segment.error = 'Please select either Distribution or Customer'
    dv_segment.errorTitle = 'Invalid Segment'

    ws.add_data_validation(dv_segment)
    dv_segment.add('H2:H1000')  # Column H for Segment

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=device_template.xlsx'

    
    # Save workbook to response
    wb.save(response)
    
    return response

# Keep the old download_template function for backward compatibility
def download_template(request):
    # Create an in-memory output file
    output = BytesIO()
    
    # Create a new workbook and add a worksheet
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()
    
    # Add some formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#0d6efd',
        'font_color': 'white',
        'border': 1
    })
    
    required_format = workbook.add_format({
        'bg_color': '#ffebee',
        'border': 1
    })
    
    optional_format = workbook.add_format({
        'bg_color': '#f5f5f5',
        'border': 1
    })
    
    # Define headers - using exact lowercase names required by the upload function
    headers = ['ip', 'hostname', 'username', 'password', 'vendor', 'ssh port', 'api port', 'segment', 'device category']
    
    # Write headers
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
    
    # Add example data
    example_data = [
        ['192.168.1.1', 'Router-Core', 'admin', 'secret123', 'Mikrotik', 22, 8728, 'Distribution', 'Router End Point'],
        ['192.168.1.2', 'Switch-Access', 'admin', 'secret123', 'Cisco', 22, 8728, 'Distribution', 'Access Switch'],
        ['10.0.0.10', 'Backup-RTR', 'admin', 'secret123', 'Juniper', 22, 8728, 'Customer', 'Router Failover'],
        ['172.16.0.1', 'BTS-Radio', 'admin', 'secret123', 'Cambium', 22, 8728, 'Distribution', 'Radio BTS'],
        ['172.16.0.2', 'Station-Radio', 'admin', 'secret123', 'Ubiquiti', 22, 8728, 'Customer', 'Radio Station']
    ]
    
    # Write example data
    for row, data in enumerate(example_data, start=1):
        for col, value in enumerate(data):
            format = required_format if col < 5 else optional_format
            worksheet.write(row, col, value, format)
    
    # Set column widths
    worksheet.set_column('A:A', 15)  # IP
    worksheet.set_column('B:B', 30)  # Hostname
    worksheet.set_column('C:C', 15)  # Username
    worksheet.set_column('D:D', 15)  # Password
    worksheet.set_column('E:E', 15)  # Vendor
    worksheet.set_column('F:G', 10)  # Ports
    worksheet.set_column('H:H', 15)  # Segment
    worksheet.set_column('I:I', 20)  # Device Category
    
    # Add data validation for Vendor column
    worksheet.data_validation('E2:E1048576', {
        'validate': 'list',
        'source': ['Mikrotik', 'Cisco', 'Juniper', 'Cambium', 'Ubiquiti'],
        'error_message': 'Please select a valid vendor'
    })
    
    # Add data validation for port numbers
    port_validation = {
        'validate': 'integer',
        'criteria': 'between',
        'minimum': 1,
        'maximum': 65535,
        'error_message': 'Port must be between 1 and 65535'
    }
    worksheet.data_validation('F2:F1048576', port_validation)
    worksheet.data_validation('G2:G1048576', port_validation)
    
    # Add data validation for Segment column
    worksheet.data_validation('H2:H1048576', {
        'validate': 'list',
        'source': ['Distribution', 'Customer'],
        'error_message': 'Please select either Distribution or Customer'
    })
    
    # Add data validation for Device Category column
    from NA_Apps.models import Device
    # Use display names (second item in tuple) instead of internal values
    device_categories = [choice[1] for choice in Device.DEVICE_CATEGORY_CHOICES]
    
    worksheet.data_validation('I2:I1048576', {
        'validate': 'list',
        'source': device_categories,
        'error_message': 'Please select a valid device category'
    })
    
    workbook.close()
    
    # Seek to start of file
    output.seek(0)
    
    # Create the HttpResponse object with Excel mime type
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=device_template.xlsx'
    
    return response

def export_logs_to_excel(logs, filename):
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    # Add formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#0d6efd',
        'font_color': 'white',
        'border': 1
    })

    # Define headers
    headers = ['Target', 'Hostname', 'Action', 'Status', 'Time', 'Messages', 'Command']
    
    # Write headers
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    # Write data
    for row, log in enumerate(logs, start=1):
        worksheet.write(row, 0, log.target)
        worksheet.write(row, 1, getattr(log, 'hostname', '-'))
        worksheet.write(row, 2, log.action)
        worksheet.write(row, 3, log.status)
        worksheet.write(row, 4, timezone.localtime(log.time).strftime('%Y-%m-%d %H:%M:%S'))
        worksheet.write(row, 5, log.messages)
        worksheet.write(row, 6, log.command if log.command else '')

    # Set column widths
    worksheet.set_column('A:A', 15)  # Target
    worksheet.set_column('B:B', 30)  # Hostname
    worksheet.set_column('C:C', 15)  # Action
    worksheet.set_column('D:D', 10)  # Status
    worksheet.set_column('E:E', 20)  # Time
    worksheet.set_column('F:F', 40)  # Messages
    worksheet.set_column('G:G', 40)  # Command

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return response

@login_required(login_url='login')
@require_POST
def filter_devices(request):
    """Handle AJAX filtering of devices"""
    try:
        data = json.loads(request.body)
        devices = Device.objects.all()
        
        # Apply filters
        device_category = data.get('device_category')
        segment = data.get('segment')
        
        if device_category:
            devices = devices.filter(device_category=device_category)
        if segment:
            devices = devices.filter(segmentation__segmentation_type=segment)
            
        # Render only the table rows
        html = render_to_string('devices_table_rows.html', {
            'all_device': devices
        }, request=request)
        
        return JsonResponse({
            'success': True,
            'html': html
        })
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=400)

@login_required(login_url='login')
def export_logs(request):
    """Export logs to Excel with hostname and respecting active filters"""
    try:
        # Get filter parameters from URL
        hostname_filter = request.GET.get('hostname', '').strip()
        status_filter = request.GET.get('status', '').strip()
        
        # Start with base queryset
        logs = Log.objects.all()
        
        # Apply filters strictly
        if hostname_filter:
            logs = logs.filter(target__icontains=hostname_filter)
        if status_filter:
            # Ensure exact status match
            logs = logs.filter(status__exact=status_filter)
        
        # Order by time
        logs = logs.order_by('-time')
        
        # Get all devices for hostname lookup with case-insensitive IP matching
        devices = {}
        for device in Device.objects.all():
            devices[device.ip_address.lower()] = device.hostname
        
        # Create an in-memory output file
        output = BytesIO()
        
        # Create a new workbook and add a worksheet
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet('Logs')
        
        # Add formats
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#0d6efd',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        success_format = workbook.add_format({
            'bg_color': '#d4edda',
            'font_color': '#155724',
            'border': 1,
            'text_wrap': True,
            'valign': 'top'
        })
        
        error_format = workbook.add_format({
            'bg_color': '#f8d7da',
            'font_color': '#721c24',
            'border': 1,
            'text_wrap': True,
            'valign': 'top'
        })
        
        date_format = workbook.add_format({
            'num_format': 'yyyy-mm-dd hh:mm:ss',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        # Define headers
        headers = ['Time', 'Target IP', 'Hostname', 'Action', 'Status', 'Messages', 'Command', 'Output']
        
        # Write headers
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        # Write data
        for row, log in enumerate(logs, start=1):
            row_format = success_format if log.status == 'Success' else error_format
            
            # Convert time to Excel datetime
            time_value = log.time.replace(tzinfo=None)
            
            # Get hostname for the IP (case-insensitive)
            hostname = devices.get(log.target.lower(), 'N/A')
            
            # Write row data
            worksheet.write(row, 0, time_value, date_format)
            worksheet.write(row, 1, log.target, row_format)
            worksheet.write(row, 2, hostname, row_format)
            worksheet.write(row, 3, log.action, row_format)
            worksheet.write(row, 4, log.status, row_format)
            worksheet.write(row, 5, log.messages, row_format)
            worksheet.write(row, 6, log.command or '', row_format)
            worksheet.write(row, 7, log.output or '', row_format)
        
        # Set column widths
        worksheet.set_column('A:A', 20)  # Time
        worksheet.set_column('B:B', 15)  # Target IP
        worksheet.set_column('C:C', 25)  # Hostname
        worksheet.set_column('D:D', 15)  # Action
        worksheet.set_column('E:E', 10)  # Status
        worksheet.set_column('F:F', 40)  # Messages
        worksheet.set_column('G:G', 40)  # Command
        worksheet.set_column('H:H', 50)  # Output
        
        # Add filter summary sheet
        summary_sheet = workbook.add_worksheet('Filter Info')
        summary_format = workbook.add_format({
            'bold': True,
            'border': 1,
            'align': 'left',
            'valign': 'vcenter'
        })
        
        # Write filter information
        summary_sheet.write(0, 0, 'Export Summary', header_format)
        summary_sheet.write(1, 0, 'Total Records:', summary_format)
        summary_sheet.write(1, 1, logs.count())
        
        row = 2
        if hostname_filter or status_filter:
            summary_sheet.write(row, 0, 'Applied Filters:', summary_format)
            row += 1
            if hostname_filter:
                summary_sheet.write(row, 0, 'Hostname/IP:', summary_format)
                summary_sheet.write(row, 1, hostname_filter)
                row += 1
            if status_filter:
                summary_sheet.write(row, 0, 'Status:', summary_format)
                summary_sheet.write(row, 1, status_filter)
        
        summary_sheet.set_column('A:A', 15)
        summary_sheet.set_column('B:B', 30)
        
        # Close the workbook
        workbook.close()
        
        # Create the HttpResponse
        output.seek(0)
        
        # Add filter info to filename if filters are active
        filename_parts = ['logs_export']
        if hostname_filter:
            filename_parts.append(f"host_{hostname_filter}")
        if status_filter:
            filename_parts.append(f"status_{status_filter}")
        filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
        
        filename = f"{'_'.join(filename_parts)}.xlsx"
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to export logs: {str(e)}'
        }, status=500)
