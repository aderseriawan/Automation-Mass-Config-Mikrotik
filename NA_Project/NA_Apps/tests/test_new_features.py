from django.urls import reverse
from django.test import Client, TestCase
from django.contrib.auth.models import User
from openpyxl import load_workbook
from io import BytesIO

from NA_Apps.models import Device, Log
from NA_Apps.views import download_device_template
from django.utils import timezone

class TestTemplateDownload(TestCase):
    """Tests for the Excel template download functionality."""

    def setUp(self):
        # Create a test user
        self.user = User.objects.create_user(
            username='testuser',
            password='testpassword'
        )
        # Create a test client and log in
        self.client = Client()
        self.client.login(username='testuser', password='testpassword')

    def test_template_has_single_column(self):
        """Test that the download_template view returns XLSX with a single column 'Device Category'"""
        # Get the template download URL
        response = self.client.get(reverse('download_template'))
        
        # Check response status and content type
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Load the Excel file and verify it has only one column with the correct header
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Check the column header
        self.assertEqual(ws['A1'].value, 'Device Category')
        
        # Verify no other columns exist
        self.assertIsNone(ws['B1'].value)


class TestDeviceCategoryFilter(TestCase):
    """Tests for device_category filtering in configure and verify views."""
    
    def setUp(self):
        # Create a test user
        self.user = User.objects.create_user(
            username='testuser',
            password='testpassword'
        )
        
        # Create test devices with different categories and unique IP addresses
        categories = [
            'router_end_point',
            'router_failover',
            'radio_bts',
            'radio_station',
            'router_bridging'
        ]
        
        self.devices = []
        for i, category in enumerate(categories):
            self.devices.append(
                Device.objects.create(
                    ip_address=f'192.168.1.{i+1}',  # Unique IP address
                    hostname=f'router-{i}',
                    username='admin',
                    password='password',
                    device_category=category
                )
            )
        
        # Create a test client and log in
        self.client = Client()
        self.client.login(username='testuser', password='testpassword')
    
    def test_configure_view_filter_by_device_category(self):
        """Test that the configure view filters devices by device_category"""
        # Test filtering by router_end_point
        response = self.client.get(reverse('configure') + '?device_category=router_end_point')
        self.assertEqual(response.status_code, 200)
        
        # Context should contain only devices with category router_end_point
        devices_in_context = response.context['devices']
        self.assertEqual(len(devices_in_context), 1)
        self.assertEqual(devices_in_context[0].device_category, 'router_end_point')
        
        # Test filtering by router_failover
        response = self.client.get(reverse('configure') + '?device_category=router_failover')
        self.assertEqual(response.status_code, 200)
        
        # Context should contain only devices with category router_failover
        devices_in_context = response.context['devices']
        self.assertEqual(len(devices_in_context), 1)
        self.assertEqual(devices_in_context[0].device_category, 'router_failover')


class TestLogFunctionality(TestCase):
    """Tests for the log view and related functionality."""
    
    def setUp(self):
        # Create a test user
        self.user = User.objects.create_user(
            username='testuser',
            password='testpassword'
        )
        
        # Create a test device
        self.device = Device.objects.create(
            ip_address='192.168.1.1',
            hostname='test-router',
            username='admin',
            password='password',
        )
        
        # Create test logs
        self.logs = [
            # Success logs
            Log.objects.create(
                target='192.168.1.1',
                action='Configure',
                status='Success',
                time=timezone.now(),
                messages=f'Success message {i}',
            )
            for i in range(3)
        ] + [
            # Error logs
            Log.objects.create(
                target='192.168.1.1',
                action='Configure',
                status='Error',
                time=timezone.now(),
                messages=f'Error message {i}',
            )
            for i in range(2)
        ]
        
        # Create a test client and log in
        self.client = Client()
        self.client.login(username='testuser', password='testpassword')
    
    def test_log_summary_counts(self):
        """Test that the log view shows accurate success/failed counts when filtering"""
        # Get log view without filters - should show all logs
        response = self.client.get(reverse('log'))
        self.assertEqual(response.status_code, 200)
        
        # Check summary counts in context
        self.assertEqual(response.context['success_count'], 3)
        self.assertEqual(response.context['failed_count'], 2)
        
        # Test filtering by success status
        response = self.client.get(reverse('log') + '?status=success')
        self.assertEqual(response.status_code, 200)
        
        # Check summary counts in context - should only show success logs
        self.assertEqual(response.context['success_count'], 3)
        self.assertEqual(response.context['failed_count'], 0)
        
        # Test filtering by failed status
        response = self.client.get(reverse('log') + '?status=failed')
        self.assertEqual(response.status_code, 200)
        
        # Check summary counts in context - should only show error logs
        self.assertEqual(response.context['success_count'], 0)
        self.assertEqual(response.context['failed_count'], 2)
    
    def test_clear_logs(self):
        """Test that the clear_logs view empties the log table and returns status 200"""
        # Verify logs exist before clearing
        self.assertEqual(Log.objects.count(), 5)
        
        # Send POST request to clear logs
        response = self.client.post(reverse('clear_logs'))
        
        # Check response status and content
        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(response.content, {'ok': 1})
        
        # Verify logs were deleted
        self.assertEqual(Log.objects.count(), 0)
